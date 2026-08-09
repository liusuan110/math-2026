import os
import math
import shutil
import numpy as np
from scipy.optimize import differential_evolution
from openpyxl import load_workbook, Workbook


# ============================================================
# 是否重新优化
# ============================================================
# False：直接使用已经搜索到的一组较优策略，速度快，结果稳定
# True ：重新运行差分进化优化，耗时较长，结果可能略有波动
RUN_OPTIMIZATION = False


# ============================================================
# 文件路径
# ============================================================

TEMPLATE_PATH = "result1.xlsx"
OUTPUT_DIR = "../output/data"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "result3.xlsx")


# ============================================================
# 基本参数
# ============================================================

g = 9.8

missile_speed = 300.0

smoke_radius = 10.0
smoke_sink_speed = 3.0
smoke_valid_time = 20.0

uav_v_min = 70.0
uav_v_max = 140.0

# M1 初始位置
M1_0 = np.array([20000.0, 0.0, 2000.0])

# FY1 初始位置
FY1_0 = np.array([17800.0, 0.0, 1800.0])

# 假目标
fake_target = np.array([0.0, 0.0, 0.0])

# 真目标圆柱体参数
target_bottom_center = np.array([0.0, 200.0, 0.0])
target_radius = 7.0
target_height = 10.0

# M1 飞行方向：直指假目标
missile_dir = (fake_target - M1_0) / np.linalg.norm(fake_target - M1_0)
missile_flight_time = np.linalg.norm(fake_target - M1_0) / missile_speed


# ============================================================
# 已搜索到的一组较优策略
# ============================================================
# 决策变量：
# x = [
#   theta, v,
#   t_drop_1, gap_12, gap_23,
#   tau_1, tau_2, tau_3
# ]
#
# theta   : FY1 航向角，弧度，以 x 轴正向为 0，逆时针为正
# v       : FY1 速度
# t_drop1 : 第 1 枚弹投放时刻
# gap_12  : 第 2 枚弹比第 1 枚弹晚投放的时间
# gap_23  : 第 3 枚弹比第 2 枚弹晚投放的时间
# tau_i   : 第 i 枚弹从投放到起爆的延时

KNOWN_GOOD_X = np.array([
    math.radians(179.5260634301327),
    138.72391855,
    0.51234277,
    2.98161327,
    1.68465937,
    4.06231020,
    5.19781994,
    5.90707697
])


# ============================================================
# 目标采样
# ============================================================

def generate_target_points(n_theta=72, n_z=11):
    """
    对真目标圆柱体侧面进行采样。
    判据：若烟幕球遮挡导弹到所有采样点的视线，则认为目标被有效遮蔽。
    """
    points = []
    cx, cy, cz = target_bottom_center

    for z in np.linspace(0.0, target_height, n_z):
        for theta in np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False):
            x = cx + target_radius * np.cos(theta)
            y = cy + target_radius * np.sin(theta)
            points.append([x, y, cz + z])

    # 加入圆柱轴线上几个点
    points.append([cx, cy, cz])
    points.append([cx, cy, cz + target_height / 2.0])
    points.append([cx, cy, cz + target_height])

    return np.array(points, dtype=float)


TARGET_POINTS_OPT = generate_target_points(n_theta=36, n_z=7)
TARGET_POINTS_FINAL = generate_target_points(n_theta=72, n_z=11)


# ============================================================
# 基础轨迹函数
# ============================================================

def missile_position(t):
    """M1 在时刻 t 的位置。"""
    return M1_0 + missile_speed * t * missile_dir


def unpack_strategy(x):
    """
    将优化变量展开为实际策略。
    """
    theta = x[0]
    v = x[1]

    t_drop_1 = x[2]
    gap_12 = x[3]
    gap_23 = x[4]

    tau_1 = x[5]
    tau_2 = x[6]
    tau_3 = x[7]

    t_drops = np.array([
        t_drop_1,
        t_drop_1 + gap_12,
        t_drop_1 + gap_12 + gap_23
    ])

    taus = np.array([tau_1, tau_2, tau_3])
    t_explodes = t_drops + taus

    uav_velocity = np.array([
        v * np.cos(theta),
        v * np.sin(theta),
        0.0
    ])

    drop_points = []
    explode_points = []

    for t_drop, tau, t_explode in zip(t_drops, taus, t_explodes):
        drop_point = FY1_0 + uav_velocity * t_drop

        explode_point = (
            FY1_0
            + uav_velocity * t_explode
            + np.array([0.0, 0.0, -0.5 * g * tau ** 2])
        )

        drop_points.append(drop_point)
        explode_points.append(explode_point)

    return {
        "theta": theta,
        "v": v,
        "uav_velocity": uav_velocity,
        "t_drops": t_drops,
        "taus": taus,
        "t_explodes": t_explodes,
        "drop_points": np.array(drop_points),
        "explode_points": np.array(explode_points)
    }


def is_feasible(x):
    """
    检查策略是否满足基本物理约束。
    """
    theta, v = x[0], x[1]

    if not (uav_v_min <= v <= uav_v_max):
        return False

    t_drop_1 = x[2]
    gap_12 = x[3]
    gap_23 = x[4]
    tau_1 = x[5]
    tau_2 = x[6]
    tau_3 = x[7]

    if t_drop_1 < 0:
        return False

    if gap_12 < 1.0 or gap_23 < 1.0:
        return False

    if tau_1 < 0 or tau_2 < 0 or tau_3 < 0:
        return False

    info = unpack_strategy(x)

    for t_explode in info["t_explodes"]:
        if t_explode < 0 or t_explode > missile_flight_time:
            return False

    for p in info["explode_points"]:
        if p[2] <= 0:
            return False

    return True


def smoke_center(t, bomb_index, x):
    """
    第 bomb_index 枚烟幕弹在时刻 t 的云团中心。
    """
    info = unpack_strategy(x)

    t_explode = info["t_explodes"][bomb_index]
    explode_point = info["explode_points"][bomb_index]

    return explode_point + np.array([
        0.0,
        0.0,
        -smoke_sink_speed * (t - t_explode)
    ])


# ============================================================
# 遮蔽判据
# ============================================================

def max_distance_to_sight_lines(t, bomb_index, x, target_points):
    """
    计算烟幕中心到“导弹-目标采样点”所有视线段的最大距离。
    最大距离 <= 10 m，则认为整个采样目标被遮蔽。
    """
    m_pos = missile_position(t)
    s_pos = smoke_center(t, bomb_index, x)

    A = m_pos
    B = target_points
    P = s_pos

    AB = B - A
    AP = P - A

    denom = np.sum(AB * AB, axis=1)
    lam = np.sum(AP * AB, axis=1) / denom
    lam = np.clip(lam, 0.0, 1.0)

    closest = A + lam[:, None] * AB
    distances = np.linalg.norm(P - closest, axis=1)

    return np.max(distances)


def is_blocked_by_bomb(t, bomb_index, x, target_points):
    """
    判断某一时刻是否被第 bomb_index 枚烟幕弹有效遮蔽。
    """
    info = unpack_strategy(x)
    t_explode = info["t_explodes"][bomb_index]

    if t < t_explode:
        return False

    if t > t_explode + smoke_valid_time:
        return False

    if t > missile_flight_time:
        return False

    d_max = max_distance_to_sight_lines(t, bomb_index, x, target_points)
    return d_max <= smoke_radius


def is_blocked_by_any_bomb(t, x, target_points):
    """
    判断某一时刻是否被任意一枚烟幕弹遮蔽。
    """
    for j in range(3):
        if is_blocked_by_bomb(t, j, x, target_points):
            return True
    return False


# ============================================================
# 时间区间计算
# ============================================================

def get_intervals_for_bomb(x, bomb_index, target_points, dt=0.01):
    """
    获取单枚烟幕弹的有效遮蔽区间。
    """
    if not is_feasible(x):
        return []

    info = unpack_strategy(x)
    t_explode = info["t_explodes"][bomb_index]

    t_start = t_explode
    t_end = min(t_explode + smoke_valid_time, missile_flight_time)

    if t_end <= t_start:
        return []

    times = np.arange(t_start, t_end + dt, dt)
    blocked = np.array([
        is_blocked_by_bomb(t, bomb_index, x, target_points)
        for t in times
    ])

    intervals = []
    inside = False
    start = None

    for i, t in enumerate(times):
        if blocked[i] and not inside:
            start = t
            inside = True

        if inside and ((not blocked[i]) or i == len(times) - 1):
            end = times[i - 1] if not blocked[i] else t
            intervals.append((start, end))
            inside = False

    return intervals


def get_union_intervals(x, target_points, dt=0.01):
    """
    获取三枚烟幕弹的联合有效遮蔽区间。
    """
    if not is_feasible(x):
        return []

    info = unpack_strategy(x)
    t_start = np.min(info["t_explodes"])
    t_end = min(np.max(info["t_explodes"]) + smoke_valid_time, missile_flight_time)

    if t_end <= t_start:
        return []

    times = np.arange(t_start, t_end + dt, dt)
    blocked = np.array([
        is_blocked_by_any_bomb(t, x, target_points)
        for t in times
    ])

    intervals = []
    inside = False
    start = None

    for i, t in enumerate(times):
        if blocked[i] and not inside:
            start = t
            inside = True

        if inside and ((not blocked[i]) or i == len(times) - 1):
            end = times[i - 1] if not blocked[i] else t
            intervals.append((start, end))
            inside = False

    return intervals


def interval_total_length(intervals):
    return sum(b - a for a, b in intervals)


def union_duration_scan(x, target_points, dt=0.02):
    """
    优化阶段使用的快速联合遮蔽时长估算。
    """
    intervals = get_union_intervals(x, target_points, dt=dt)
    return interval_total_length(intervals)


# ============================================================
# 优化
# ============================================================

def objective(x):
    """
    差分进化最小化目标函数。
    因为要最大化遮蔽时间，所以返回负数。
    """
    if not is_feasible(x):
        return 1e6

    duration = union_duration_scan(x, TARGET_POINTS_OPT, dt=0.03)
    return -duration


def run_optimization():
    """
    重新搜索问题3策略。
    注意：这是非凸优化，结果和随机种子、采样密度有关。
    """
    bounds = [
        (math.radians(160), math.radians(200)),  # theta
        (70.0, 140.0),                           # v
        (0.0, 6.0),                              # t_drop_1
        (1.0, 6.0),                              # gap_12
        (1.0, 6.0),                              # gap_23
        (0.0, 8.0),                              # tau_1
        (0.0, 8.0),                              # tau_2
        (0.0, 8.0)                               # tau_3
    ]

    result = differential_evolution(
        objective,
        bounds=bounds,
        maxiter=80,
        popsize=14,
        tol=1e-3,
        seed=2025,
        polish=False,
        workers=1,
        updating="immediate"
    )

    return result.x


# ============================================================
# 写入 Excel
# ============================================================

def prepare_workbook():
    """
    优先使用官方模板。
    如果模板不存在，则创建一个同结构工作簿。
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(TEMPLATE_PATH):
        shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
        wb = load_workbook(OUTPUT_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = [
            "无人机运动方向",
            "无人机运动速度 (m/s)",
            "烟幕干扰弹编号",
            "烟幕干扰弹投放点的x坐标 (m)",
            "烟幕干扰弹投放点的y坐标 (m)",
            "烟幕干扰弹投放点的z坐标 (m)",
            "烟幕干扰弹起爆点的x坐标 (m)",
            "烟幕干扰弹起爆点的y坐标 (m)",
            "烟幕干扰弹起爆点的z坐标 (m)",
            "有效干扰时长 (s)"
        ]

        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)

        for i in range(3):
            ws.cell(row=2 + i, column=3, value=i + 1)

        ws.cell(
            row=6,
            column=1,
            value="注：以x轴为正向，逆时针方向为正，取值0~360（度）。"
        )

    return wb, ws


def write_result_to_excel(x, output_path=OUTPUT_PATH):
    """
    将策略结果写入 result1.xlsx。
    """
    info = unpack_strategy(x)

    theta_deg = math.degrees(info["theta"]) % 360.0
    v = info["v"]

    drop_points = info["drop_points"]
    explode_points = info["explode_points"]

    # 单枚烟幕弹有效时长
    individual_durations = []
    individual_intervals = []

    for j in range(3):
        intervals = get_intervals_for_bomb(
            x,
            j,
            TARGET_POINTS_FINAL,
            dt=0.01
        )
        individual_intervals.append(intervals)
        individual_durations.append(interval_total_length(intervals))

    wb, ws = prepare_workbook()

    for j in range(3):
        row = 2 + j

        ws.cell(row=row, column=1, value=theta_deg)
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=3, value=j + 1)

        ws.cell(row=row, column=4, value=float(drop_points[j, 0]))
        ws.cell(row=row, column=5, value=float(drop_points[j, 1]))
        ws.cell(row=row, column=6, value=float(drop_points[j, 2]))

        ws.cell(row=row, column=7, value=float(explode_points[j, 0]))
        ws.cell(row=row, column=8, value=float(explode_points[j, 1]))
        ws.cell(row=row, column=9, value=float(explode_points[j, 2]))

        ws.cell(row=row, column=10, value=float(individual_durations[j]))

    # 简单设置列宽，避免本地打开时太挤
    widths = {
        "A": 16,
        "B": 18,
        "C": 16,
        "D": 24,
        "E": 24,
        "F": 24,
        "G": 24,
        "H": 24,
        "I": 24,
        "J": 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)

    return {
        "theta_deg": theta_deg,
        "v": v,
        "individual_durations": individual_durations,
        "individual_intervals": individual_intervals,
        "union_intervals": get_union_intervals(x, TARGET_POINTS_FINAL, dt=0.01),
        "drop_points": drop_points,
        "explode_points": explode_points,
        "t_drops": info["t_drops"],
        "t_explodes": info["t_explodes"],
        "taus": info["taus"],
        "output_path": output_path
    }


# ============================================================
# 主程序
# ============================================================

def main():
    if RUN_OPTIMIZATION:
        print("开始重新优化，可能需要较长时间...")
        x_opt = run_optimization()

        # 为防止重新优化结果不如已知较优解，二者都计算，取更好者
        candidates = [KNOWN_GOOD_X, x_opt]

        best_x = None
        best_duration = -1.0

        for x in candidates:
            duration = union_duration_scan(
                x,
                TARGET_POINTS_FINAL,
                dt=0.01
            )

            if duration > best_duration:
                best_duration = duration
                best_x = x

        x = best_x

    else:
        x = KNOWN_GOOD_X

    result = write_result_to_excel(x)

    print("\n问题3计算完成")
    print("=" * 60)

    print(f"FY1 飞行方向角: {result['theta_deg']:.6f} deg")
    print(f"FY1 飞行速度: {result['v']:.6f} m/s")

    print("\n投放与起爆时刻:")
    for i in range(3):
        print(
            f"第{i + 1}枚: "
            f"t_drop = {result['t_drops'][i]:.6f} s, "
            f"tau = {result['taus'][i]:.6f} s, "
            f"t_explode = {result['t_explodes'][i]:.6f} s"
        )

    print("\n投放点:")
    for i, p in enumerate(result["drop_points"], start=1):
        print(f"第{i}枚: ({p[0]:.6f}, {p[1]:.6f}, {p[2]:.6f})")

    print("\n起爆点:")
    for i, p in enumerate(result["explode_points"], start=1):
        print(f"第{i}枚: ({p[0]:.6f}, {p[1]:.6f}, {p[2]:.6f})")

    print("\n单枚烟幕弹有效遮蔽区间:")
    for i, intervals in enumerate(result["individual_intervals"], start=1):
        duration = interval_total_length(intervals)

        if len(intervals) == 0:
            print(f"第{i}枚: 无有效遮蔽，时长 0.000000 s")
        else:
            interval_text = ", ".join(
                [f"[{a:.6f}, {b:.6f}]" for a, b in intervals]
            )
            print(f"第{i}枚: {interval_text}, 合计 {duration:.6f} s")

    union_duration = interval_total_length(result["union_intervals"])

    print("\n三枚烟幕弹联合有效遮蔽区间:")
    for a, b in result["union_intervals"]:
        print(f"[{a:.6f}, {b:.6f}]，持续 {b - a:.6f} s")

    print(f"\n联合有效遮蔽总时长: {union_duration:.6f} s")

    print(f"\n结果已保存到: {result['output_path']}")


if __name__ == "__main__":
    main()