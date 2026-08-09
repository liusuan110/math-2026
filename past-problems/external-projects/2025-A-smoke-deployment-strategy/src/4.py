import os
import math
import shutil
import numpy as np
from scipy.optimize import differential_evolution
from openpyxl import load_workbook, Workbook


# ============================================================
# 问题4：FY1、FY2、FY3各投放1枚烟幕弹，干扰M1
# 输出文件：../output/data/result4.xlsx
# ============================================================

RUN_OPTIMIZATION = False
# False：直接使用一组已搜索到的策略，运行快，结果稳定
# True ：重新搜索，运行时间较长，结果可能有小幅波动


# ============================================================
# 文件路径
# ============================================================

TEMPLATE_PATH = "result2.xlsx"
OUTPUT_DIR = "../output/data"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "result4.xlsx")
#(1)

# ============================================================
# 基本参数
# ============================================================

g = 9.8

missile_speed = 300.0

smoke_radius = 10.0
smoke_sink_speed = 3.0
smoke_valid_time = 20.0

uav_v_min = 70.0
uav_v_max = 140.0

# M1 初始位置
M1_0 = np.array([20000.0, 0.0, 2000.0])

# 假目标
fake_target = np.array([0.0, 0.0, 0.0])

# FY1、FY2、FY3 初始位置
UAV_NAMES = ["FY1", "FY2", "FY3"]
UAV_0 = [
    np.array([17800.0, 0.0, 1800.0]),
    np.array([12000.0, 1400.0, 1400.0]),
    np.array([6000.0, -3000.0, 700.0])
]

# 真目标圆柱体参数
target_bottom_center = np.array([0.0, 200.0, 0.0])
target_radius = 7.0
target_height = 10.0

# M1 飞行方向：直指假目标
missile_dir = (fake_target - M1_0) / np.linalg.norm(fake_target - M1_0)
missile_flight_time = np.linalg.norm(fake_target - M1_0) / missile_speed


# ============================================================
# 已搜索到的一组可行较优策略
# ============================================================

KNOWN_GOOD_X = np.array([
    math.radians(5.26149853491536),
    124.53864067834121,
    0.7436191127240832,
    0.21394945688628608,

    math.radians(296.83690007213164),
    128.6771504609418,
    6.8672525925629015,
    4.994461745197506,

    math.radians(75.06399818264352),
    123.94943830029652,
    24.210274795862134,
    1.5551214644316547
])


# ============================================================
# 目标采样
# ============================================================

def generate_target_points(n_theta=72, n_z=11):
    """
    对真目标圆柱体表面采样。
    判据：烟幕球遮挡导弹到全部采样点的视线，则认为目标被有效遮蔽。
    """
    points = []
    cx, cy, cz = target_bottom_center

    for z in np.linspace(0.0, target_height, n_z):
        for theta in np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False):
            x = cx + target_radius * np.cos(theta)
            y = cy + target_radius * np.sin(theta)
            points.append([x, y, cz + z])

    points.append([cx, cy, cz])
    points.append([cx, cy, cz + target_height / 2.0])
    points.append([cx, cy, cz + target_height])

    return np.array(points, dtype=float)


# 优化时用较粗采样，最终校验用较细采样
TARGET_POINTS_OPT = generate_target_points(n_theta=18, n_z=5)
TARGET_POINTS_FINAL = generate_target_points(n_theta=72, n_z=11)


# ============================================================
# 基础轨迹函数
# ============================================================

def missile_position(t):
    """M1 在时刻 t 的位置。"""
    return M1_0 + missile_speed * t * missile_dir


def unpack_one_uav(y, uav_index):
    """
    展开单架无人机的一枚烟幕弹策略。
    y = [theta, v, t_drop, tau]
    """
    theta, v, t_drop, tau = y

    uav_velocity = np.array([
        v * np.cos(theta),
        v * np.sin(theta),
        0.0
    ])

    t_explode = t_drop + tau

    drop_point = UAV_0[uav_index] + uav_velocity * t_drop

    explode_point = (
        UAV_0[uav_index]
        + uav_velocity * t_explode
        + np.array([0.0, 0.0, -0.5 * g * tau ** 2])
    )

    return {
        "theta": theta,
        "v": v,
        "t_drop": t_drop,
        "tau": tau,
        "t_explode": t_explode,
        "uav_velocity": uav_velocity,
        "drop_point": drop_point,
        "explode_point": explode_point
    }


def unpack_all(x):
    """展开三架无人机的全部策略。"""
    info = []

    for i in range(3):
        y = x[4 * i: 4 * i + 4]
        info.append(unpack_one_uav(y, i))

    return info


def is_feasible_one(y, uav_index):
    """检查单架无人机策略是否满足物理约束。"""
    theta, v, t_drop, tau = y

    if not (uav_v_min <= v <= uav_v_max):
        return False

    if t_drop < 0.0:
        return False

    if tau < 0.0:
        return False

    info = unpack_one_uav(y, uav_index)

    if info["t_explode"] < 0.0:
        return False

    if info["t_explode"] > missile_flight_time:
        return False

    if info["explode_point"][2] <= 0.0:
        return False

    return True


def is_feasible_all(x):
    """检查三架无人机策略是否全部可行。"""
    for i in range(3):
        y = x[4 * i: 4 * i + 4]
        if not is_feasible_one(y, i):
            return False
    return True


def smoke_center(t, x, uav_index):
    """第 uav_index 架无人机烟幕云团中心在时刻 t 的位置。"""
    y = x[4 * uav_index: 4 * uav_index + 4]
    info = unpack_one_uav(y, uav_index)

    return info["explode_point"] + np.array([
        0.0,
        0.0,
        -smoke_sink_speed * (t - info["t_explode"])
    ])


# ============================================================
# 遮蔽判据
# ============================================================

def max_distance_to_sight_lines(t, x, uav_index, target_points):
    """
    计算烟幕中心到“导弹-目标采样点”所有视线段的最大距离。
    如果最大距离 <= smoke_radius，则认为该烟幕弹可遮蔽整个采样目标。
    """
    m_pos = missile_position(t)
    s_pos = smoke_center(t, x, uav_index)

    A = m_pos
    B = target_points
    P = s_pos

    AB = B - A
    AP = P - A

    denom = np.sum(AB * AB, axis=1)
    lam = np.sum(AP * AB, axis=1) / denom
    lam = np.clip(lam, 0.0, 1.0)

    closest = A + lam[:, None] * AB
    distances = np.linalg.norm(P - closest, axis=1)

    return np.max(distances)


def is_blocked_by_uav(t, x, uav_index, target_points):
    """判断时刻 t 是否被某架无人机释放的烟幕弹有效遮蔽。"""
    y = x[4 * uav_index: 4 * uav_index + 4]

    if not is_feasible_one(y, uav_index):
        return False

    info = unpack_one_uav(y, uav_index)
    t_explode = info["t_explode"]

    if t < t_explode:
        return False

    if t > t_explode + smoke_valid_time:
        return False

    if t > missile_flight_time:
        return False

    d_max = max_distance_to_sight_lines(t, x, uav_index, target_points)

    return d_max <= smoke_radius


def is_blocked_by_any(t, x, target_points):
    """判断时刻 t 是否被任意一枚烟幕弹有效遮蔽。"""
    for i in range(3):
        if is_blocked_by_uav(t, x, i, target_points):
            return True
    return False


# ============================================================
# 遮蔽区间计算
# ============================================================

def get_intervals_for_uav(x, uav_index, target_points, dt=0.005):
    """
    获取某一枚烟幕弹的有效遮蔽区间。
    """
    y = x[4 * uav_index: 4 * uav_index + 4]

    if not is_feasible_one(y, uav_index):
        return []

    info = unpack_one_uav(y, uav_index)

    t_start = info["t_explode"]
    t_end = min(info["t_explode"] + smoke_valid_time, missile_flight_time)

    if t_end <= t_start:
        return []

    times = np.arange(t_start, t_end + dt, dt)
    blocked = np.array([
        is_blocked_by_uav(t, x, uav_index, target_points)
        for t in times
    ])

    intervals = []
    inside = False
    start = None

    for i, t in enumerate(times):
        if blocked[i] and not inside:
            start = t
            inside = True

        if inside and ((not blocked[i]) or i == len(times) - 1):
            end = times[i - 1] if not blocked[i] else t
            intervals.append((start, end))
            inside = False

    return intervals


def get_union_intervals(x, target_points, dt=0.005):
    """
    获取三枚烟幕弹的联合有效遮蔽区间。
    """
    if not is_feasible_all(x):
        return []

    info_all = unpack_all(x)

    t_start = min(info["t_explode"] for info in info_all)
    t_end = min(
        max(info["t_explode"] for info in info_all) + smoke_valid_time,
        missile_flight_time
    )

    if t_end <= t_start:
        return []

    times = np.arange(t_start, t_end + dt, dt)
    blocked = np.array([
        is_blocked_by_any(t, x, target_points)
        for t in times
    ])

    intervals = []
    inside = False
    start = None

    for i, t in enumerate(times):
        if blocked[i] and not inside:
            start = t
            inside = True

        if inside and ((not blocked[i]) or i == len(times) - 1):
            end = times[i - 1] if not blocked[i] else t
            intervals.append((start, end))
            inside = False

    return intervals


def interval_total_length(intervals):
    """计算区间总长度。"""
    return sum(b - a for a, b in intervals)


# ============================================================
# 重新优化
# ============================================================

def objective_at_target_time(y, uav_index, target_time):
    """
    重新优化时使用的单机目标函数。
    三架无人机分别瞄准早期、中期、后期三个遮蔽时间段。
    """
    if not is_feasible_one(y, uav_index):
        return 1e6

    info = unpack_one_uav(y, uav_index)

    if not (info["t_explode"] <= target_time <= info["t_explode"] + smoke_valid_time):
        return 1e5 + abs(target_time - info["t_explode"])

    x_temp = KNOWN_GOOD_X.copy()
    x_temp[4 * uav_index: 4 * uav_index + 4] = y

    return max_distance_to_sight_lines(
        target_time,
        x_temp,
        uav_index,
        TARGET_POINTS_OPT
    )


def optimize_one_uav(uav_index, target_time, bounds, seed):
    """优化单架无人机的一枚烟幕弹策略。"""
    result = differential_evolution(
        lambda y: objective_at_target_time(y, uav_index, target_time),
        bounds=bounds,
        maxiter=100,
        popsize=12,
        tol=1e-5,
        seed=seed,
        polish=True,
        workers=1
    )

    return result.x


def run_optimization():
    """
    重新搜索第4问策略。
    """
    bounds_list = [
        [
            (math.radians(-20.0), math.radians(30.0)),
            (70.0, 140.0),
            (0.0, 5.0),
            (0.0, 8.0)
        ],
        [
            (math.radians(240.0), math.radians(320.0)),
            (70.0, 140.0),
            (0.0, 20.0),
            (0.0, 15.0)
        ],
        [
            (math.radians(40.0), math.radians(120.0)),
            (70.0, 140.0),
            (0.0, 40.0),
            (0.0, 15.0)
        ]
    ]

    target_times = [3.5, 14.5, 27.0]

    pieces = []

    for i in range(3):
        print(f"正在优化 {UAV_NAMES[i]} ...")
        y = optimize_one_uav(
            uav_index=i,
            target_time=target_times[i],
            bounds=bounds_list[i],
            seed=2025 + i
        )
        pieces.append(y)

    return np.concatenate(pieces)


# ============================================================
# Excel 写入
# ============================================================

def prepare_workbook():

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(TEMPLATE_PATH):
        shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
        wb = load_workbook(OUTPUT_PATH)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "无人机编号",
        "无人机运动方向",
        "无人机运动速度 (m/s)",
        "烟幕干扰弹投放点的x坐标 (m)",
        "烟幕干扰弹投放点的y坐标 (m)",
        "烟幕干扰弹投放点的z坐标 (m)",
        "烟幕干扰弹起爆点的x坐标 (m)",
        "烟幕干扰弹起爆点的y坐标 (m)",
        "烟幕干扰弹起爆点的z坐标 (m)",
        "有效干扰时长 (s)"
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for i, name in enumerate(UAV_NAMES, start=2):
        ws.cell(row=i, column=1, value=name)

    ws.cell(
        row=6,
        column=1,
        value="注：无人机运动方向以x轴正方向为0度，逆时针为正。"
    )

    return wb, ws


def find_header_columns(ws):
    """
    尝试根据表头自动识别列。
    如果识别失败，使用默认列号。
    """
    default_cols = {
        "uav": 1,
        "direction": 2,
        "speed": 3,
        "drop_x": 4,
        "drop_y": 5,
        "drop_z": 6,
        "explode_x": 7,
        "explode_y": 8,
        "explode_z": 9,
        "duration": 10
    }

    header_row = 1
    max_col = ws.max_column

    headers = {}
    for col in range(1, max_col + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None:
            headers[col] = str(value)

    if len(headers) == 0:
        return default_cols

    cols = default_cols.copy()

    for col, text in headers.items():
        if "无人机编号" in text or text.strip() == "无人机":
            cols["uav"] = col
        elif "方向" in text:
            cols["direction"] = col
        elif "速度" in text:
            cols["speed"] = col
        elif "投放点" in text and "x" in text.lower():
            cols["drop_x"] = col
        elif "投放点" in text and "y" in text.lower():
            cols["drop_y"] = col
        elif "投放点" in text and "z" in text.lower():
            cols["drop_z"] = col
        elif "起爆点" in text and "x" in text.lower():
            cols["explode_x"] = col
        elif "起爆点" in text and "y" in text.lower():
            cols["explode_y"] = col
        elif "起爆点" in text and "z" in text.lower():
            cols["explode_z"] = col
        elif "有效" in text and "时长" in text:
            cols["duration"] = col

    return cols


def write_result_to_excel(x):
    """
    将第4问结果写入 result4.xlsx。
    """
    wb, ws = prepare_workbook()
    cols = find_header_columns(ws)

    info_all = unpack_all(x)

    for i in range(3):
        row = 2 + i
        info = info_all[i]

        theta_deg = math.degrees(info["theta"]) % 360.0

        intervals = get_intervals_for_uav(
            x,
            i,
            TARGET_POINTS_FINAL,
            dt=0.005
        )
        duration = interval_total_length(intervals)

        drop = info["drop_point"]
        explode = info["explode_point"]

        ws.cell(row=row, column=cols["uav"], value=UAV_NAMES[i])
        ws.cell(row=row, column=cols["direction"], value=float(theta_deg))
        ws.cell(row=row, column=cols["speed"], value=float(info["v"]))

        ws.cell(row=row, column=cols["drop_x"], value=float(drop[0]))
        ws.cell(row=row, column=cols["drop_y"], value=float(drop[1]))
        ws.cell(row=row, column=cols["drop_z"], value=float(drop[2]))

        ws.cell(row=row, column=cols["explode_x"], value=float(explode[0]))
        ws.cell(row=row, column=cols["explode_y"], value=float(explode[1]))
        ws.cell(row=row, column=cols["explode_z"], value=float(explode[2]))

        ws.cell(row=row, column=cols["duration"], value=float(duration))

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    wb.save(OUTPUT_PATH)


# ============================================================
# 打印结果
# ============================================================

def print_result(x):
    info_all = unpack_all(x)

    print("\n问题4计算结果")
    print("=" * 70)

    for i in range(3):
        info = info_all[i]
        theta_deg = math.degrees(info["theta"]) % 360.0

        intervals = get_intervals_for_uav(
            x,
            i,
            TARGET_POINTS_FINAL,
            dt=0.005
        )
        duration = interval_total_length(intervals)

        print(f"\n{UAV_NAMES[i]}")
        print(f"飞行方向角: {theta_deg:.6f} deg")
        print(f"飞行速度: {info['v']:.6f} m/s")
        print(f"投放时刻: {info['t_drop']:.6f} s")
        print(f"起爆延时: {info['tau']:.6f} s")
        print(f"起爆时刻: {info['t_explode']:.6f} s")

        drop = info["drop_point"]
        explode = info["explode_point"]

        print(f"投放点: ({drop[0]:.6f}, {drop[1]:.6f}, {drop[2]:.6f})")
        print(f"起爆点: ({explode[0]:.6f}, {explode[1]:.6f}, {explode[2]:.6f})")

        if len(intervals) == 0:
            print("有效遮蔽区间: 无")
        else:
            text = ", ".join([f"[{a:.6f}, {b:.6f}]" for a, b in intervals])
            print(f"有效遮蔽区间: {text}")

        print(f"单弹有效遮蔽时长: {duration:.6f} s")

    union_intervals = get_union_intervals(
        x,
        TARGET_POINTS_FINAL,
        dt=0.005
    )
    union_duration = interval_total_length(union_intervals)

    print("\n三枚烟幕弹联合有效遮蔽区间:")
    for a, b in union_intervals:
        print(f"[{a:.6f}, {b:.6f}]，持续 {b - a:.6f} s")

    print(f"\n联合有效遮蔽总时长: {union_duration:.6f} s")
    print(f"\n结果文件已保存到: {OUTPUT_PATH}")


# ============================================================
# 主程序
# ============================================================

def main():
    if RUN_OPTIMIZATION:
        x_new = run_optimization()

        old_duration = interval_total_length(
            get_union_intervals(KNOWN_GOOD_X, TARGET_POINTS_FINAL, dt=0.01)
        )
        new_duration = interval_total_length(
            get_union_intervals(x_new, TARGET_POINTS_FINAL, dt=0.01)
        )

        if new_duration > old_duration:
            x = x_new
        else:
            x = KNOWN_GOOD_X
    else:
        x = KNOWN_GOOD_X

    write_result_to_excel(x)
    print_result(x)


if __name__ == "__main__":
    main()
