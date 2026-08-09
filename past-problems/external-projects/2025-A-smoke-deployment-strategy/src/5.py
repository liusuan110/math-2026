import os
import math
import shutil
import numpy as np
from openpyxl import Workbook, load_workbook


# ============================================================
# 问题5：5架无人机，每架最多3枚烟幕弹，干扰M1、M2、M3
# 输出文件：../output/data/result5.xlsx
# ============================================================
#
# 本脚本默认使用一组已搜索到的可行策略，并重新按几何遮蔽模型计算时长。
# 如需重新局部优化，可把 RUN_LOCAL_OPTIMIZATION 改为 True。
#
# 依赖：
# pip install numpy openpyxl scipy
#
# 若 RUN_LOCAL_OPTIMIZATION=False，则不需要 scipy。
# 若 RUN_LOCAL_OPTIMIZATION=True，则需要 scipy。
#
# 输出：
# ../output/data/result5.xlsx
#
# 说明：
# 1. 同一无人机最多投放3枚烟幕弹。
# 2. 同一无人机相邻投放时间间隔必须 >= 1s。
# 3. 每架无人机航向、速度一旦确定后不再调整。
# 4. 烟幕弹脱离无人机后保持无人机水平速度，竖直方向自由落体。
# 5. 起爆后云团中心以3m/s匀速下沉，有效半径10m，有效时间20s。
# 6. 遮蔽判据：烟幕球遮挡导弹到真目标圆柱体采样点的全部视线。


# ============================================================
# 开关
# ============================================================

RUN_LOCAL_OPTIMIZATION = False
# False：直接复现当前策略，速度快、结果稳定。
# True ：在当前策略附近做局部数值搜索，耗时较长，结果可能有波动。

EVAL_DT = 0.01
# 最终计算遮蔽区间的时间步长。
# 想更精细可改为 0.005，但会更慢。

TEMPLATE_PATH = "result3.xlsx"
OUTPUT_DIR = "../output/data"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "result5.xlsx")
#(1)

# ============================================================
# 基本参数
# ============================================================

g = 9.8

missile_speed = 300.0

smoke_radius = 10.0
smoke_sink_speed = 3.0
smoke_valid_time = 20.0

uav_v_min = 70.0
uav_v_max = 140.0

fake_target = np.array([0.0, 0.0, 0.0])

MISSILE_0 = {
    "M1": np.array([20000.0, 0.0, 2000.0]),
    "M2": np.array([19000.0, 600.0, 2100.0]),
    "M3": np.array([18000.0, -600.0, 1900.0])
}

UAV_0 = {
    "FY1": np.array([17800.0, 0.0, 1800.0]),
    "FY2": np.array([12000.0, 1400.0, 1400.0]),
    "FY3": np.array([6000.0, -3000.0, 700.0]),
    "FY4": np.array([11000.0, 2000.0, 1800.0]),
    "FY5": np.array([13000.0, -2000.0, 1300.0])
}

UAV_NAMES = ["FY1", "FY2", "FY3", "FY4", "FY5"]
MISSILE_NAMES = ["M1", "M2", "M3"]

# 真目标：下底面圆心为(0, 200, 0)，半径7m，高10m
target_bottom_center = np.array([0.0, 200.0, 0.0])
target_radius = 7.0
target_height = 10.0


# ============================================================
# 当前保存策略
# ============================================================
# 每架无人机：
#   heading_deg: 无人机运动方向角，单位度，以x轴正向为0，逆时针为正。
#   speed: 无人机速度，m/s。
#   bombs: 每枚烟幕弹的投放时刻、起爆延时、干扰导弹编号。
#
# 注意：
# 这里保存的是策略变量，不是简单照抄Excel坐标。
# 运行时会根据运动方程重新计算投放点、起爆点、遮蔽区间和有效时长。

SAVED_STRATEGY = {
    "FY1": {
        "heading_deg": 179.526063,
        "speed": 138.723919,
        "bombs": [
            {"missile": "M1", "t_drop": 0.5123427669746268, "tau": 4.06231018540544},
            {"missile": "M1", "t_drop": 3.4939560263707987, "tau": 5.197819923746203},
            {"missile": "M1", "t_drop": 5.178615390288446, "tau": 5.907076955526204},
        ],
    },
    "FY2": {
        "heading_deg": 293.500119,
        "speed": 118.170029,
        "bombs": [
            {"missile": "M2", "t_drop": 6.681896880667671, "tau": 2.2031733363796997},
            {"missile": "M2", "t_drop": 7.682022593686839, "tau": 1.3172406832947443},
            {"missile": "M2", "t_drop": 9.070492795425173, "tau": 0.06301405421743311},
        ],
    },
    "FY3": {
        "heading_deg": 84.928250,
        "speed": 137.223220,
        "bombs": [
            {"missile": "M3", "t_drop": 18.72331783591857, "tau": 2.2695951634647713},
            {"missile": "M3", "t_drop": 19.72331783493165, "tau": 1.1524528188181975},
            {"missile": "M3", "t_drop": 20.723317834588958, "tau": 0.0},
        ],
    },
    "FY4": {
        "heading_deg": 265.000957,
        "speed": 136.726473,
        "bombs": [
            {"missile": "M2", "t_drop": 0.7840116809817113, "tau": 10.988850623616154},
            {"missile": "M1", "t_drop": 2.3104839582656624, "tau": 11.991929514395792},
            {"missile": "M3", "t_drop": 5.882729493930647, "tau": 11.435490723308217},
        ],
    },
    "FY5": {
        "heading_deg": 121.060350,
        "speed": 140.000000,
        "bombs": [
            {"missile": "M3", "t_drop": 11.143828918436922, "tau": 2.3414353702595587},
            {"missile": "M3", "t_drop": 12.143828915748362, "tau": 1.2689082732267742},
            {"missile": "M3", "t_drop": 13.240278337854823, "tau": 0.09205857213958524},
        ],
    },
}


# ============================================================
# 目标采样
# ============================================================

def generate_target_points(n_theta=72, n_z=11):
    """
    生成真目标圆柱体的表面采样点。
    """
    points = []
    cx, cy, cz = target_bottom_center

    for z in np.linspace(0.0, target_height, n_z):
        for theta in np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False):
            x = cx + target_radius * np.cos(theta)
            y = cy + target_radius * np.sin(theta)
            points.append([x, y, cz + z])

    points.append([cx, cy, cz])
    points.append([cx, cy, cz + target_height / 2.0])
    points.append([cx, cy, cz + target_height])

    return np.array(points, dtype=float)


TARGET_POINTS_FINAL = generate_target_points(n_theta=72, n_z=11)
TARGET_POINTS_OPT = generate_target_points(n_theta=24, n_z=7)


# ============================================================
# 基础运动函数
# ============================================================

def unit(v):
    n = np.linalg.norm(v)
    if n == 0:
        raise ValueError("零向量无法归一化。")
    return v / n


MISSILE_DIR = {
    name: unit(fake_target - pos)
    for name, pos in MISSILE_0.items()
}

MISSILE_FLIGHT_TIME = {
    name: np.linalg.norm(fake_target - pos) / missile_speed
    for name, pos in MISSILE_0.items()
}


def missile_position(missile_name, t):
    """
    导弹在时刻t的位置。
    """
    return MISSILE_0[missile_name] + missile_speed * t * MISSILE_DIR[missile_name]


def uav_direction_vector(heading_deg):
    """
    由方向角得到水平单位方向向量。
    """
    theta = math.radians(heading_deg)
    return np.array([math.cos(theta), math.sin(theta), 0.0])


def get_drop_point(uav_name, heading_deg, speed, t_drop):
    """
    无人机等高度匀速直线飞行时，烟幕弹投放点。
    """
    direction = uav_direction_vector(heading_deg)
    return UAV_0[uav_name] + speed * t_drop * direction


def get_explode_point(uav_name, heading_deg, speed, t_drop, tau):
    """
    烟幕弹起爆点。
    t_drop: 投放时刻。
    tau: 投放后到起爆的延时。
    """
    direction = uav_direction_vector(heading_deg)
    t_explode = t_drop + tau

    horizontal_pos = UAV_0[uav_name] + speed * t_explode * direction
    vertical_drop = np.array([0.0, 0.0, -0.5 * g * tau ** 2])

    return horizontal_pos + vertical_drop


def smoke_center(explode_point, t_explode, t):
    """
    起爆后烟幕云团中心。
    """
    return explode_point + np.array([
        0.0,
        0.0,
        -smoke_sink_speed * (t - t_explode)
    ])


# ============================================================
# 策略展开与可行性检查
# ============================================================

def strategy_to_records(strategy):
    """
    将策略字典展开为15条烟幕弹记录。
    """
    records = []

    for uav_name in UAV_NAMES:
        uav_data = strategy[uav_name]
        heading_deg = float(uav_data["heading_deg"])
        speed = float(uav_data["speed"])

        for bomb_idx, bomb in enumerate(uav_data["bombs"], start=1):
            t_drop = float(bomb["t_drop"])
            tau = float(bomb["tau"])
            missile_name = bomb["missile"]
            t_explode = t_drop + tau

            drop_point = get_drop_point(
                uav_name,
                heading_deg,
                speed,
                t_drop
            )

            explode_point = get_explode_point(
                uav_name,
                heading_deg,
                speed,
                t_drop,
                tau
            )

            records.append({
                "uav": uav_name,
                "heading_deg": heading_deg,
                "speed": speed,
                "bomb_idx": bomb_idx,
                "missile": missile_name,
                "t_drop": t_drop,
                "tau": tau,
                "t_explode": t_explode,
                "drop_point": drop_point,
                "explode_point": explode_point,
            })

    return records


def check_strategy(strategy, verbose=True):
    """
    检查策略是否满足速度、投放间隔、高度、导弹飞行时间等约束。
    """
    ok = True

    for uav_name in UAV_NAMES:
        data = strategy[uav_name]
        speed = data["speed"]

        if not (uav_v_min <= speed <= uav_v_max):
            ok = False
            if verbose:
                print(f"[约束错误] {uav_name} 速度超限: {speed}")

        drops = [bomb["t_drop"] for bomb in data["bombs"]]

        for i, t in enumerate(drops):
            if t < 0:
                ok = False
                if verbose:
                    print(f"[约束错误] {uav_name} 第{i+1}枚投放时刻为负。")

        for i in range(len(drops) - 1):
            gap = drops[i + 1] - drops[i]
            if gap < 1.0 - 1e-8:
                ok = False
                if verbose:
                    print(f"[约束错误] {uav_name} 第{i+1}和第{i+2}枚投放间隔不足1s: {gap}")

    records = strategy_to_records(strategy)

    for r in records:
        if r["tau"] < -1e-8:
            ok = False
            if verbose:
                print(f"[约束错误] {r['uav']} 第{r['bomb_idx']}枚起爆延时为负。")

        if r["explode_point"][2] < -1e-8:
            ok = False
            if verbose:
                print(f"[约束错误] {r['uav']} 第{r['bomb_idx']}枚起爆点高度小于0。")

        flight_time = MISSILE_FLIGHT_TIME[r["missile"]]
        if r["t_explode"] > flight_time:
            ok = False
            if verbose:
                print(f"[约束错误] {r['uav']} 第{r['bomb_idx']}枚起爆晚于导弹到达假目标。")

    return ok


# ============================================================
# 遮蔽判据
# ============================================================

def max_distance_to_sight_lines(missile_name, t, smoke_pos, target_points):
    """
    烟幕中心到“导弹-目标采样点”所有视线段的最大距离。
    若最大距离 <= smoke_radius，则认为该烟幕球遮蔽整个采样目标。
    """
    m_pos = missile_position(missile_name, t)

    A = m_pos
    B = target_points
    P = smoke_pos

    AB = B - A
    AP = P - A

    denom = np.sum(AB * AB, axis=1)
    lam = np.sum(AP * AB, axis=1) / denom
    lam = np.clip(lam, 0.0, 1.0)

    closest = A + lam[:, None] * AB
    distances = np.linalg.norm(P - closest, axis=1)

    return float(np.max(distances))


def is_record_blocking(record, t, target_points):
    """
    判断某一枚烟幕弹在时刻t是否对其指定导弹有效遮蔽。
    """
    missile_name = record["missile"]

    if t < record["t_explode"]:
        return False

    if t > record["t_explode"] + smoke_valid_time:
        return False

    if t > MISSILE_FLIGHT_TIME[missile_name]:
        return False

    s_pos = smoke_center(
        record["explode_point"],
        record["t_explode"],
        t
    )

    d_max = max_distance_to_sight_lines(
        missile_name,
        t,
        s_pos,
        target_points
    )

    return d_max <= smoke_radius


def get_intervals_for_record(record, target_points, dt=0.01):
    """
    计算单枚烟幕弹对其指定导弹的有效遮蔽区间。
    """
    t_start = record["t_explode"]
    t_end = min(
        record["t_explode"] + smoke_valid_time,
        MISSILE_FLIGHT_TIME[record["missile"]]
    )

    if t_end <= t_start:
        return []

    times = np.arange(t_start, t_end + dt, dt)
    flags = np.array([
        is_record_blocking(record, t, target_points)
        for t in times
    ])

    return flags_to_intervals(times, flags)


def flags_to_intervals(times, flags):
    """
    将布尔遮蔽序列转换为区间。
    """
    intervals = []
    inside = False
    start = None

    for i, t in enumerate(times):
        if flags[i] and not inside:
            start = t
            inside = True

        if inside and ((not flags[i]) or i == len(times) - 1):
            end = times[i - 1] if not flags[i] else t
            intervals.append((float(start), float(end)))
            inside = False

    return intervals


def merge_intervals(intervals):
    """
    合并区间，避免重复计算重叠遮蔽时间。
    """
    if len(intervals) == 0:
        return []

    intervals = sorted(intervals, key=lambda x: x[0])
    merged = [list(intervals[0])]

    for a, b in intervals[1:]:
        if a <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], b)
        else:
            merged.append([a, b])

    return [(float(a), float(b)) for a, b in merged]


def interval_total_length(intervals):
    return float(sum(b - a for a, b in intervals))


def evaluate_strategy(strategy, target_points=TARGET_POINTS_FINAL, dt=0.01):
    """
    计算：
    1. 单枚烟幕弹遮蔽区间与时长；
    2. 各导弹合并后的联合遮蔽区间与时长；
    3. 三枚导弹联合遮蔽总时长。
    """
    records = strategy_to_records(strategy)

    for r in records:
        intervals = get_intervals_for_record(r, target_points, dt=dt)
        r["intervals"] = intervals
        r["duration"] = interval_total_length(intervals)

    missile_intervals = {name: [] for name in MISSILE_NAMES}

    for r in records:
        missile_intervals[r["missile"]].extend(r["intervals"])

    missile_merged = {
        name: merge_intervals(missile_intervals[name])
        for name in MISSILE_NAMES
    }

    missile_duration = {
        name: interval_total_length(missile_merged[name])
        for name in MISSILE_NAMES
    }

    single_sum = sum(r["duration"] for r in records)
    merged_total = sum(missile_duration.values())

    return {
        "records": records,
        "missile_merged": missile_merged,
        "missile_duration": missile_duration,
        "single_sum": single_sum,
        "merged_total": merged_total,
    }


# ============================================================
# 可选：局部优化
# ============================================================

def strategy_to_vector(strategy):
    """
    将策略字典转为优化向量。
    每架无人机8个变量：
    [heading_deg, speed, t_drop1, gap12, gap23, tau1, tau2, tau3]
    合计40维。
    """
    xs = []

    for uav_name in UAV_NAMES:
        data = strategy[uav_name]
        bombs = data["bombs"]

        t1 = bombs[0]["t_drop"]
        t2 = bombs[1]["t_drop"]
        t3 = bombs[2]["t_drop"]

        xs.extend([
            data["heading_deg"],
            data["speed"],
            t1,
            t2 - t1,
            t3 - t2,
            bombs[0]["tau"],
            bombs[1]["tau"],
            bombs[2]["tau"],
        ])

    return np.array(xs, dtype=float)


def vector_to_strategy(x, template_strategy):
    """
    将优化向量转回策略字典。
    导弹分配保持template_strategy不变。
    """
    strategy = {}

    for i, uav_name in enumerate(UAV_NAMES):
        base = 8 * i

        heading_deg = float(x[base + 0]) % 360.0
        speed = float(x[base + 1])

        t1 = float(x[base + 2])
        gap12 = float(x[base + 3])
        gap23 = float(x[base + 4])

        tau1 = float(x[base + 5])
        tau2 = float(x[base + 6])
        tau3 = float(x[base + 7])

        t_drops = [t1, t1 + gap12, t1 + gap12 + gap23]
        taus = [tau1, tau2, tau3]

        bombs = []

        for j in range(3):
            missile_name = template_strategy[uav_name]["bombs"][j]["missile"]

            bombs.append({
                "missile": missile_name,
                "t_drop": t_drops[j],
                "tau": taus[j],
            })

        strategy[uav_name] = {
            "heading_deg": heading_deg,
            "speed": speed,
            "bombs": bombs,
        }

    return strategy


def local_optimization(initial_strategy):
    """
    在当前策略附近做局部优化。
    目标函数：最大化按导弹合并后的联合遮蔽时长。
    """
    from scipy.optimize import differential_evolution

    x0 = strategy_to_vector(initial_strategy)

    bounds = []

    for i, uav_name in enumerate(UAV_NAMES):
        base = 8 * i
        heading0 = x0[base + 0]
        speed0 = x0[base + 1]
        t10 = x0[base + 2]
        gap120 = x0[base + 3]
        gap230 = x0[base + 4]
        tau10 = x0[base + 5]
        tau20 = x0[base + 6]
        tau30 = x0[base + 7]

        bounds.extend([
            (heading0 - 8.0, heading0 + 8.0),
            (max(uav_v_min, speed0 - 15.0), min(uav_v_max, speed0 + 15.0)),
            (max(0.0, t10 - 3.0), t10 + 3.0),
            (1.0, max(1.0, gap120 + 3.0)),
            (1.0, max(1.0, gap230 + 3.0)),
            (max(0.0, tau10 - 3.0), tau10 + 3.0),
            (max(0.0, tau20 - 3.0), tau20 + 3.0),
            (max(0.0, tau30 - 3.0), tau30 + 3.0),
        ])

    def objective(x):
        strategy = vector_to_strategy(x, initial_strategy)

        if not check_strategy(strategy, verbose=False):
            return 1e6

        # 优化阶段使用较粗目标采样和时间步长，加快速度。
        result = evaluate_strategy(
            strategy,
            target_points=TARGET_POINTS_OPT,
            dt=0.04
        )

        return -result["merged_total"]

    result = differential_evolution(
        objective,
        bounds=bounds,
        maxiter=40,
        popsize=8,
        tol=1e-3,
        seed=2025,
        polish=False,
        workers=1,
        updating="immediate"
    )

    candidate = vector_to_strategy(result.x, initial_strategy)

    old_result = evaluate_strategy(
        initial_strategy,
        target_points=TARGET_POINTS_FINAL,
        dt=EVAL_DT
    )
    new_result = evaluate_strategy(
        candidate,
        target_points=TARGET_POINTS_FINAL,
        dt=EVAL_DT
    )

    if new_result["merged_total"] > old_result["merged_total"]:
        return candidate

    return initial_strategy


# ============================================================
# Excel输出
# ============================================================

def prepare_workbook():

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(TEMPLATE_PATH):
        shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)
        wb = load_workbook(OUTPUT_PATH)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "无人机编号",
        "无人机运动方向",
        "无人机运动速度 (m/s)",
        "烟幕干扰弹编号",
        "烟幕干扰弹投放点的x坐标 (m)",
        "烟幕干扰弹投放点的y坐标 (m)",
        "烟幕干扰弹投放点的z坐标 (m)",
        "烟幕干扰弹起爆点的x坐标 (m)",
        "烟幕干扰弹起爆点的y坐标 (m)",
        "烟幕干扰弹起爆点的z坐标 (m)",
        "有效干扰时长 (s)",
        "干扰的导弹编号"
    ]

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    return wb, ws


def write_result_to_excel(eval_result):
    """
    将问题5结果写入 result5.xlsx。
    """
    wb, ws = prepare_workbook()

    headers = [
        "无人机编号",
        "无人机运动方向",
        "无人机运动速度 (m/s)",
        "烟幕干扰弹编号",
        "烟幕干扰弹投放点的x坐标 (m)",
        "烟幕干扰弹投放点的y坐标 (m)",
        "烟幕干扰弹投放点的z坐标 (m)",
        "烟幕干扰弹起爆点的x坐标 (m)",
        "烟幕干扰弹起爆点的y坐标 (m)",
        "烟幕干扰弹起爆点的z坐标 (m)",
        "有效干扰时长 (s)",
        "干扰的导弹编号"
    ]

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    records = eval_result["records"]

    for row_idx, r in enumerate(records, start=2):
        drop = r["drop_point"]
        explode = r["explode_point"]

        ws.cell(row=row_idx, column=1, value=r["uav"])
        ws.cell(row=row_idx, column=2, value=round(r["heading_deg"] % 360.0, 6))
        ws.cell(row=row_idx, column=3, value=round(r["speed"], 6))
        ws.cell(row=row_idx, column=4, value=r["bomb_idx"])

        ws.cell(row=row_idx, column=5, value=round(float(drop[0]), 6))
        ws.cell(row=row_idx, column=6, value=round(float(drop[1]), 6))
        ws.cell(row=row_idx, column=7, value=round(float(drop[2]), 6))

        ws.cell(row=row_idx, column=8, value=round(float(explode[0]), 6))
        ws.cell(row=row_idx, column=9, value=round(float(explode[1]), 6))
        ws.cell(row=row_idx, column=10, value=round(float(explode[2]), 6))

        ws.cell(row=row_idx, column=11, value=round(float(r["duration"]), 6))
        ws.cell(row=row_idx, column=12, value=r["missile"])

    note_row = 18
    ws.cell(row=note_row, column=2, value="注：以x轴为正向，逆时针方向为正，取值0~360（度）。")

    summary_row = 20
    ws.cell(row=summary_row, column=1, value="结果摘要")
    ws.cell(row=summary_row + 1, column=1, value="指标")
    ws.cell(row=summary_row + 1, column=2, value="数值/说明")

    ws.cell(row=summary_row + 2, column=1, value="单弹有效干扰时长合计 (s)")
    ws.cell(row=summary_row + 2, column=2, value=round(eval_result["single_sum"], 6))

    ws.cell(row=summary_row + 3, column=1, value="按导弹合并后的联合有效干扰总时长 (s)")
    ws.cell(row=summary_row + 3, column=2, value=round(eval_result["merged_total"], 6))

    for k, missile_name in enumerate(MISSILE_NAMES, start=4):
        intervals = eval_result["missile_merged"][missile_name]
        text = "; ".join([f"[{a:.3f}, {b:.3f}]" for a, b in intervals])

        ws.cell(row=summary_row + k, column=1, value=f"{missile_name}联合遮蔽区间")
        ws.cell(row=summary_row + k, column=2, value=text)

    ws.cell(
        row=summary_row + 7,
        column=1,
        value="说明"
    )
    ws.cell(
        row=summary_row + 7,
        column=2,
        value="各无人机最多投放3枚；同一无人机投放间隔均>=1s；方向角以x轴正向逆时针计。"
    )

    for col in range(1, 13):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    wb.save(OUTPUT_PATH)


# ============================================================
# 结果打印
# ============================================================

def print_result(eval_result):
    records = eval_result["records"]

    print("\n问题5计算完成")
    print("=" * 80)

    print("\n烟幕弹策略：")
    for r in records:
        drop = r["drop_point"]
        explode = r["explode_point"]

        print(
            f"{r['uav']} 第{r['bomb_idx']}枚 -> {r['missile']} | "
            f"方向={r['heading_deg'] % 360.0:.6f} deg, "
            f"速度={r['speed']:.6f} m/s, "
            f"t_drop={r['t_drop']:.6f} s, "
            f"tau={r['tau']:.6f} s, "
            f"t_explode={r['t_explode']:.6f} s, "
            f"有效时长={r['duration']:.6f} s"
        )

        print(
            f"    投放点=({drop[0]:.6f}, {drop[1]:.6f}, {drop[2]:.6f}), "
            f"起爆点=({explode[0]:.6f}, {explode[1]:.6f}, {explode[2]:.6f})"
        )

    print("\n按导弹合并后的遮蔽区间：")
    for missile_name in MISSILE_NAMES:
        intervals = eval_result["missile_merged"][missile_name]
        duration = eval_result["missile_duration"][missile_name]

        if intervals:
            text = "; ".join([f"[{a:.6f}, {b:.6f}]" for a, b in intervals])
        else:
            text = "无"

        print(f"{missile_name}: {text}，合计 {duration:.6f} s")

    print(f"\n单弹有效干扰时长合计: {eval_result['single_sum']:.6f} s")
    print(f"按导弹合并后的联合有效干扰总时长: {eval_result['merged_total']:.6f} s")
    print(f"\n结果文件已保存到: {OUTPUT_PATH}")


# ============================================================
# 主程序
# ============================================================

def main():
    strategy = SAVED_STRATEGY

    if not check_strategy(strategy, verbose=True):
        raise RuntimeError("当前策略不满足基本约束，请检查SAVED_STRATEGY。")

    if RUN_LOCAL_OPTIMIZATION:
        print("开始局部优化，可能耗时较长...")
        strategy = local_optimization(strategy)

    eval_result = evaluate_strategy(
        strategy,
        target_points=TARGET_POINTS_FINAL,
        dt=EVAL_DT
    )

    write_result_to_excel(eval_result)
    print_result(eval_result)


if __name__ == "__main__":
    main()
